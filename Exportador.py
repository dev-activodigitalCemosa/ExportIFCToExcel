import ifcopenshell
import ifcopenshell.util
import ifcopenshell.geom
import ifcopenshell.util.element
import ifcopenshell.util.placement
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import re
from tqdm import trange
from time import sleep
import customtkinter as ctk
from tkinter import Toplevel, Scrollbar, IntVar, Checkbutton, filedialog, messagebox
import threading
import logging

# CONFIGURACIÓN DE CustomTkinter
ctk.set_appearance_mode("System")  # "System", "Dark", "Light"
ctk.set_default_color_theme("blue")  # "blue", "green", "dark-blue"

# FUNCIONES ORIGINALES
entity_properties = []

def get_entities_filtered(ifcschema_entities, get_types):
    # ents_not_collect = ["IfcGeometricRepresentationItem", "IfcObject",
    #                     "IfcObjectDefinition", "IfcProduct", "IfcRelationship",
    #                     "IfcRepresentationItem", "IfcRoot", "IfcCarPoiLisD"]
    ents_not_collect = ["IfcElement"]

    if get_types:
        ents = [e for e in ifcschema_entities if "type" in e.lower()]
    else:
        ents = [e for e in ifcschema_entities if "type" not in e.lower() and e in ents_not_collect]
    return (len(ents), ents)

def get_entity_properties(ifc_entName):
    entities = ifc_file.by_type(ifc_entName)
    global entity_properties
    entity_properties = []
    for en in entities:
        props = ifcopenshell.util.element.get_psets(en)
        filtered_props = {key: value for key, value in props.items() if 'Pset' not in key}
        entity_properties.append(filtered_props)
    entity_properties = pd.json_normalize(entity_properties)
    return entity_properties

def get_ents_info_to_df(ifc_entName,df):
    #campos a filtrar
    fields_to_exclude = ['Description','ObjectPlacement','Representation']
    # Obtener la información básica
    entities = ifc_file.by_type(ifc_entName)
    entity_info = [en.get_info() for en in entities]
    entity_container = [ifcopenshell.util.element.get_container(en) for en in entities]
    entity_coord = [ifcopenshell.util.placement.get_local_placement(en.ObjectPlacement)[:,3][:3] for en in entities]
    
    
    # Crear DataFrames
    df_info = pd.DataFrame(entity_info)
    df_container = pd.DataFrame(entity_container, columns=['Location'])
    df_coord = pd.DataFrame(entity_coord, columns=['X', 'Y', 'Z'])  # Nombrar las columnas XYZ

     # Filtrar campos específicos en df_info
    if fields_to_exclude:
        existing_fields = [col for col in df_info.columns if col not in fields_to_exclude]
        df_info = df_info[existing_fields]

    # Concatenar los DataFrames
    combined_df = pd.concat([df_info,df, df_container, df_coord], axis=1)
    
    return combined_df

def contract_entName(entName, trunc_n=3):
    entName_split = re.findall('[A-Z][^A-Z]*', entName)
    return "".join([s[:trunc_n] if len(s) >= trunc_n else s for s in entName_split])

def create_ws(wb, ws_name):
    if ws_name in wb.sheetnames:
        ws = wb.create_sheet(title=ws_name + "_1")
    else:
        ws = wb.create_sheet(title=ws_name)
    return ws

def create_ws_and_table(wb,df, ifc_entName, table_suffix_counter={}):
    df = get_ents_info_to_df(ifc_entName,df)
    
    if len(ifc_entName) > 20:
        ifc_entName = contract_entName(ifc_entName, trunc_n=3)
    
    ws = create_ws(wb, ifc_entName)
    
    # Escribir encabezados jerárquicos
    if not df.empty:
        if isinstance(df.columns, pd.MultiIndex):
            # Escribir encabezados de nivel 1
            for col_num, (level1, _) in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_num, value=str(level1) if level1 else '')
            # Escribir encabezados de nivel 2
            for col_num, (_, level2) in enumerate(df.columns, start=1):
                ws.cell(row=2, column=col_num, value=str(level2) if level2 else '')
        else:
            # Para columnas simples, solo escribir el encabezado en la primera fila
            for col_num, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_num, value=str(col_name))

        # Escribir los datos del DataFrame
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append([str(v) if str(v).startswith('#') or '(' in str(v) else v for v in row])

        df_shape = df.shape
        last_column_letter = get_excel_column_letter(df_shape[1]-1)
        tbl_xl_rangeAddress = f"A1:{last_column_letter}{df_shape[0] + 2}"

        if not re.match(r'^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$', tbl_xl_rangeAddress):
            raise ValueError(f"Invalid range address generated: {tbl_xl_rangeAddress}")

        base_table_name = ifc_entName
        suffix = table_suffix_counter.get(base_table_name, 0)
        unique_table_name = f"{base_table_name}_{suffix}"
        table_suffix_counter[base_table_name] = suffix + 1

        tbl = Table(displayName=unique_table_name, ref=tbl_xl_rangeAddress)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    
    return True

def remove_ws(wb, ws_name="Sheet"):
    if ws_name in wb.sheetnames:
        ws_to_delete = wb[ws_name]
        wb.remove(ws_to_delete)

def purge_wb(wb):
    for ws in wb.sheetnames:
        ws = wb[ws]
    for tbl in ws.tables.values(): ws._tables.remove(tbl)
    remove_ws(wb, ws_name=ws)
    return True

def get_excel_column_letter(n):
    string = ""
    while n >= 0:
        n, remainder = divmod(n, 26)
        string = chr(65 + remainder) + string
        n -= 1
    return string

# FUNCIONES PARA LA INTERFAZ GRÁFICA

def select_ifc_file():
    global ifc_filename, ifc_basepath, ifc_file
    file_path = filedialog.askopenfilename(filetypes=[("IFC files", "*.ifc")])
    if file_path:
        ifc_filename = os.path.basename(file_path).split('.')[0]
        ifc_basepath = os.path.dirname(file_path)
        ifc_file = ifcopenshell.open(file_path)
        file_entry.delete(0, ctk.END)
        file_entry.insert(0, file_path)
        console_output.insert(ctk.END, f"Archivo IFC seleccionado: {file_path}\n")

def select_destination_folder():
    global destination_folder
    destination_folder = filedialog.askdirectory()
    folder_entry.delete(0, ctk.END)
    folder_entry.insert(0, destination_folder)
    console_output.insert(ctk.END, f"Carpeta de destino seleccionada: {destination_folder}\n")

def start_processing_thread():
    # Deshabilitar botones
    btn_select_file.configure(state=ctk.DISABLED)
    btn_select_folder.configure(state=ctk.DISABLED)
    btn_process.configure(state=ctk.DISABLED)
    # Crear un hilo separado para ejecutar el procesamiento
    thread = threading.Thread(target=process_ifc_file)
    thread.start()

logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def process_ifc_file():
    logging.info("Inicio del procesamiento")
    try:
        if not ifc_file or not destination_folder:
            messagebox.showerror("Error", "Seleccione un archivo IFC y una carpeta de destino antes de continuar.")
            return

        global ifc_getTypes
        ifc_getTypes = False

        # Carga del modelo y procesamiento
        sch_entities_names = [e.name() for e in ifcopenshell.schema_by_name(ifc_file.schema).entities()]
        entities_names_in_use = []
        for en in sch_entities_names:
            try:
                if len(ifc_file.by_type(en)) != 0:
                    entities_names_in_use.append(en)
            except RuntimeError as e:
                console_output.insert(ctk.END, f"Advertencia: {e} - Entidad '{en}' ignorada.\n")
        ents = get_entities_filtered(entities_names_in_use, get_types=ifc_getTypes)


        xls_filename = os.path.join(destination_folder, datetime.now().strftime("%Y-%m-%d") + f"_{ifc_filename}_entTypes_{str(ifc_getTypes)}_2.xlsx")
        wb = Workbook()
        purge_wb(wb)

        table_suffix_counter = {}
        t = trange(len(ents[1]), desc='Entity: ', leave=True)
        progress_bar.set(0)
        total_ents = len(ents[1])

        for i, en in zip(t, ents[1]):
            t.set_description(f"Entity: {en}")
            t.refresh()
            get_entity_properties(en)
            filtered_df = open_pset_selection()
            create_ws_and_table(wb,filtered_df, en, table_suffix_counter)
                # Calcula el progreso actual y actualiza la barra de progreso
            progress = (i + 1) / total_ents
            progress_bar.set(progress)  # Actualiza la barra de progreso
            progress_bar.update_idletasks()

        remove_ws(wb, ws_name="Sheet")
        wb.save(xls_filename)
        console_output.insert(ctk.END, f"Archivo {xls_filename} ha sido guardado exitosamente!\n")

        # Muestra la ventana emergente para selección de propiedades
        
    except Exception as e:
        logging.error(f"Error durante el procesamiento: {str(e)}")
    finally:
        enable_buttons()  # Habilitar botones al finalizar el procesamiento

def enable_buttons():
    """Función para habilitar los botones después de finalizar el procesamiento."""
    btn_select_file.configure(state=ctk.NORMAL)
    btn_select_folder.configure(state=ctk.NORMAL)
    btn_process.configure(state=ctk.NORMAL)

def open_pset_selection():
 
    # Crear la ventana emergente
    pset_window = ctk.CTkToplevel(app)
    pset_window.title("Seleccionar Propiedades (Psets)")
    pset_window.geometry("450x650")
    pset_window.attributes("-topmost", True)
    # Obtener el tamaño y posición de la ventana principal
    app_width = app.winfo_width()
    app_height = app.winfo_height()
    app_x = app.winfo_rootx()
    app_y = app.winfo_rooty()
    
    # Obtener el tamaño de la ventana emergente
    window_width = 450
    window_height = 650
    
    # Calcular la posición central
    x = app_x + (app_width - window_width) // 2
    y = app_y + (app_height - window_height) // 2
    
    # Ajustar la posición de la ventana emergente
    pset_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # Crear un frame con scroll para los checkboxes
    frame_canvas = ctk.CTkFrame(pset_window)
    frame_canvas.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = ctk.CTkCanvas(frame_canvas,bg="#2b2b2b",highlightthickness=0)  # Sin bordes y fondo igual al de la ventana
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)

    scrollbar = ctk.CTkScrollbar(frame_canvas, command=canvas.yview, orientation="vertical")
    scrollbar.pack(side="right", fill="y",padx=5, pady=5)
    
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Crear un frame para los checkboxes dentro del canvas
    checkbox_frame = ctk.CTkFrame(canvas)
    canvas.create_window((0, 0), window=checkbox_frame, anchor="nw")

    # Variable para almacenar los valores de los checkboxes
    selected_psets = {}
    global entity_properties
    # Crear un checkbox para cada `pset`
    for i, pset in enumerate(entity_properties):
        selected_psets[pset] = IntVar(value=1)  # Todos seleccionados por defecto
        checkbox = Checkbutton(checkbox_frame, text=pset, variable=selected_psets[pset],
                               onvalue=1, offvalue=0, font=("Helvetica", 12),  # Ajusta la fuente
                               bg="#2b2b2b", fg="white", selectcolor="#1a73e8", activebackground="#2b2b2b")
        checkbox.pack(anchor="w", pady=2)  # Añade espacio vertical entre checkboxes
    # Actualizar el tamaño del frame interno en función del contenido
    checkbox_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))
    filtered_df = None
    
    # Botón de confirmación
    def confirm_selection():
        nonlocal filtered_df
        selected_columns = [pset for pset, var in selected_psets.items() if var.get() == 1]
        filtered_df = entity_properties[selected_columns]
        pset_window.destroy()

    confirm_button = ctk.CTkButton(pset_window, text="Confirmar selección", command=confirm_selection,
                                   width=200, height=40, fg_color="#1a73e8", text_color="white",
                                   hover_color="#1451a3")
    confirm_button.pack(pady=15)
    app.wait_window(pset_window)
    return filtered_df


# CONFIGURACIÓN DE LA INTERFAZ GRÁFICA

app = ctk.CTk()  # Cambiar tk.Tk() por ctk.CTk()
app.title("Exportador de propiedades IFC")
app.geometry("700x500")  # Ajusta el tamaño de la ventana

# Frame para seleccionar archivo IFC
frame_file = ctk.CTkFrame(app)
frame_file.pack(fill=ctk.X, padx=20, pady=5)

file_entry = ctk.CTkEntry(frame_file)
file_entry.pack(side=ctk.LEFT, fill=ctk.X, expand=True, padx=(0, 5))

btn_select_file = ctk.CTkButton(frame_file, text="Seleccionar Archivo IFC", command=select_ifc_file, width=25)
btn_select_file.pack(side=ctk.RIGHT)

# Frame para seleccionar carpeta de destino
frame_folder = ctk.CTkFrame(app)
frame_folder.pack(fill=ctk.X, padx=20, pady=5)

folder_entry = ctk.CTkEntry(frame_folder)
folder_entry.pack(side=ctk.LEFT, fill=ctk.X, expand=True, padx=(0, 5))

btn_select_folder = ctk.CTkButton(frame_folder, text="Seleccionar Carpeta de Destino", command=select_destination_folder, width=25)
btn_select_folder.pack(side=ctk.RIGHT)

# Botón para procesar el archivo
btn_process = ctk.CTkButton(app, text="Procesar Archivo", command=start_processing_thread)
btn_process.pack(pady=10)

# Barra de progreso
progress_bar = ctk.CTkProgressBar(app, orientation="horizontal", width=100, mode="determinate")
progress_bar.pack(fill=ctk.X, padx=20, pady=10)

# Consola de salida
console_output = ctk.CTkTextbox(app, wrap=ctk.WORD, width=80, height=10)
console_output.pack(padx=20, pady=10, fill=ctk.BOTH, expand=True)

app.mainloop()
